from openpyxl import load_workbook
import shutil

# 从样板生成对应班级课表文件
def create_xlsx(class_name):
    # 路径与文件名
    example_path = r'../data/example/课表模板.xlsx'
    result_path = r'../data/result/'
    file_name = class_name + '课表.xlsx'

    if example_path.endswith('课表模板.xlsx'):
        shutil.copyfile(example_path, result_path + '/' + file_name)
        print('复制成功')
    else:
        print('未找到正确课表模板')

# 导入数据到课表文件
def excel(schedules, week_time, class_time, class_name):
    # 课表文件名
    file_name = class_name + '课表.xlsx'

    # 打开Excel文件
    wb = load_workbook('../data/result/'+file_name)
    ws = wb['Sheet0']

    # 目标单元格位置
    # 星期
    if week_time == '周一':
        column = 3
    elif week_time == '周二':
        column = 4
    elif week_time == '周三':
        column = 5
    elif week_time == '周四':
        column = 6
    elif week_time == '周五':
        column = 7
    # 课程时间
    if class_time == '1-2节':
        row = 3
    elif class_time == '3-4节':
        row = 4
    elif class_time == '5-6节':
        row = 5
    elif class_time == '7-8节':
        row = 6
    elif class_time == '9-10节':
        row = 7

    # 选择写入的课表信息及课表位置
    class_info = ''
    for schedule in schedules:
        # 检查每个属性是否为None，如果不是，则添加到class_info中
        if schedule.course_name is not None:
            class_info += schedule.course_name + '/'
        if schedule.start_end_week is not None:
            class_info += schedule.start_end_week + '/'
        if schedule.venue_requirement is not None:
            class_info += schedule.venue_requirement + '/'
        if schedule.teacher is not None:
            class_info += schedule.teacher

    # 写入单元格
    ws.cell(row, column).value = class_info

    # 保存Excel文件
    wb.save('../data/result/'+file_name)
