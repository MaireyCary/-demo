from data import custom_positions
from openpyxl import load_workbook
from countWeeks import final_schedule

# 加载模板Excel文件
file_path = r'../data/example/人员安排表.xlsx'  # 模板文件的路径
wb = load_workbook(file_path)
ws = wb.active  # 激活第一个工作表

# 将排班结果写入到对应的单元格中
for week, days in final_schedule.items():
    for day, periods in days.items():
        for period, people in periods.items():
            # 获取对应的单元格位置列表
            cell_positions = custom_positions.get(week, {}).get(day, {}).get(period, ["", ""])
            # 检查cell_positions是否有效
            if all(cell_positions):  # 确保两个单元格位置都不为空
                # 将人员名单分别写入两个单元格
                if len(people) == 2:  # 确保有两个人
                    ws[cell_positions[0]] = people[0]  # 写入第一个人
                    ws[cell_positions[1]] = people[1]  # 写入第二个人
                elif len(people) == 1:  # 如果只有一个人
                    ws[cell_positions[0]] = people[0]  # 写入第一个人
                    ws[cell_positions[1]] = ""  # 空出第二个单元格
                else:
                    print(f"提示: {day}的{period}没有人员安排.\n")

# 保存到新的Excel文件中，以免覆盖原始模板文件
output_file = r'..\data\result\人员安排表.xlsx'
wb.save(output_file)
wb.close()  # 关闭工作簿
print(f"排班结果已写入到: {output_file}")
